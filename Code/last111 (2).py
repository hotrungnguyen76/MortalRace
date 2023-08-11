import pygame, sys, random, numpy
import math
import openpyxl, time
from pygame.locals import *

pygame.init()

BLACK =(0,0,0)
YELLOW = (255,255,0)
RED = (255,0,0)
GREY = (128,128,128)
PINK =(255,123,133)
WHITE = (255,255,255)
GREEN = (0,128,0)
BLUE = (30,144,255)

FPS = 30
fpsClock = pygame.time.Clock()


WINDOWWIDTH = 1300 # Chiều dài cửa sổ
WINDOWHEIGHT = 650 # Chiều cao cửa sổ

font = pygame.font.SysFont('comicsansms',40)
screen = pygame.display.set_mode((WINDOWWIDTH,WINDOWHEIGHT))
pygame.display.set_caption('Mortal Race')

smallfont = pygame.font.SysFont('comicsansms',25)
smallfont1 = pygame.font.SysFont('comicsansms',18)
mediumfont = pygame.font.SysFont('comicsansms',30)
largefont1 = pygame.font.SysFont('gillsansultra',75)
largefont2 = pygame.font.SysFont('comicsansms',75)
largefont3 = pygame.font.SysFont('gillsansultra',75)
largefont4 = pygame.font.SysFont('gillsansultra',50)


backgroundStart = pygame.image.load('start.png')


background = [0,0,0,0,0]
background[0] = pygame.image.load('samac.png')
background[1] = pygame.image.load('space.png')
background[2] = pygame.image.load('ocean.png')
background[3] = pygame.image.load('start1.png')
background[4] = pygame.image.load('blue.png')
appleimg = pygame.image.load('apple.png')
mapstart = pygame.image.load('mapstart.png')
choose_bet= pygame.image.load('choosebet.png')


theme1 = pygame.image.load('samacmap.png')
theme2 = pygame.image.load('spacemap.png')
theme3 = pygame.image.load('oceanmap.png')


item = [0,0,0,0,0]
item[0] = pygame.image.load('speed_up.png')
item[1] = pygame.image.load('speed_down.png')
item[2] = pygame.image.load('go_back.png')
item[3] = pygame.image.load('dizzy.png')
item[4] = pygame.image.load('finish.png')

setList=[0, 0, 0, 0]
setList[0] = pygame.image.load('choose_set_1.png')
setList[1] = pygame.image.load('choose_set_2.png')
setList[2] = pygame.image.load('choose_set_3.png')
setList[3] = pygame.image.load('choose_set_4.png')

playerChoose = [0,0,0,0]
playerChoose[0]=pygame.image.load('setList1.png')
playerChoose[1]=pygame.image.load('setList2.png')
playerChoose[2]=pygame.image.load('setList3.png')
playerChoose[3]=pygame.image.load('setList4.png')

character=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
character[0]=pygame.image.load('1.png')
character[1]=pygame.image.load('2.png')
character[2]=pygame.image.load('3.png')
character[3]=pygame.image.load('4.png')
character[4]=pygame.image.load('5.png')
character[5]=pygame.image.load('6.png')
character[6]=pygame.image.load('7.png')
character[7]=pygame.image.load('8.png')
character[8]=pygame.image.load('9.png')
character[9]=pygame.image.load('10.png')
character[10]=pygame.image.load('11.png')
character[11]=pygame.image.load('12.png')
character[12]=pygame.image.load('13.png')
character[13]=pygame.image.load('14.png')
character[14]=pygame.image.load('15.png')
character[15]=pygame.image.load('16.png')

character_move=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
character_move[0]=pygame.image.load('1_move.png')
character_move[1]=pygame.image.load('2_move.png')
character_move[2]=pygame.image.load('3_move.png')
character_move[3]=pygame.image.load('4_move.png')
character_move[4]=pygame.image.load('5_move.png')
character_move[5]=pygame.image.load('6_move.png')
character_move[6]=pygame.image.load('7_move.png')
character_move[7]=pygame.image.load('8_move.png')
character_move[8]=pygame.image.load('9_move.png')
character_move[9]=pygame.image.load('10_move.png')
character_move[10]=pygame.image.load('11_move.png')
character_move[11]=pygame.image.load('12_move.png')
character_move[12]=pygame.image.load('13_move.png')
character_move[13]=pygame.image.load('14_move.png')
character_move[14]=pygame.image.load('15_move.png')
character_move[15]=pygame.image.load('16_move.png')

character_move_1=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
character_move_1[0]=pygame.image.load('1_move_1.png')
character_move_1[1]=pygame.image.load('2_move_1.png')
character_move_1[2]=pygame.image.load('3_move_1.png')
character_move_1[3]=pygame.image.load('4_move_1.png')
character_move_1[4]=pygame.image.load('5_move_1.png')
character_move_1[5]=pygame.image.load('6_move_1.png')
character_move_1[6]=pygame.image.load('7_move_1.png')
character_move_1[7]=pygame.image.load('8_move_1.png')
character_move_1[8]=pygame.image.load('9_move_1.png')
character_move_1[9]=pygame.image.load('10_move_1.png')
character_move_1[10]=pygame.image.load('11_move_1.png')
character_move_1[11]=pygame.image.load('12_move_1.png')
character_move_1[12]=pygame.image.load('13_move_1.png')
character_move_1[13]=pygame.image.load('14_move_1.png')
character_move_1[14]=pygame.image.load('15_move_1.png')
character_move_1[15]=pygame.image.load('16_move_1.png')
bat = pygame.image.load('bat.png')
tat = pygame.image.load('tat.png')

nhanh = pygame.image.load('flash.png')
cham = pygame.image.load('snail.png')
quaylai = pygame.image.load('quaylai.png')
Choang = pygame.image.load('choang.png')

sound = 1
ten=['','','','']
snakeScore = 0
snakeMoney = 0
username = ''
money_change =0 

# CÁC BIẾN CẦN RESET
player=[0,0,0,0]

itemPlayer1 = [0,0,0,0,0]
itemPlayer1[0] = random.randint(0,3)
itemPlayer1[1] = random.randint(0,3)
itemPlayer1[2] = random.randint(0,3)
itemPlayer1[3] = random.randint(0,4)
itemPlayer1[4] = random.randint(0,4)

itemPlayer2 = [0,0,0,0,0]
itemPlayer2[0] = random.randint(0,3)
itemPlayer2[1] = random.randint(0,3)
itemPlayer2[2] = random.randint(0,3)
itemPlayer2[3] = random.randint(0,4)
itemPlayer2[4] = random.randint(0,4)

itemPlayer3 = [0,0,0,0,0]
itemPlayer3[0] = random.randint(0,3)
itemPlayer3[1] = random.randint(0,3)
itemPlayer3[2] = random.randint(0,3)
itemPlayer3[3] = random.randint(0,4)
itemPlayer3[4] = random.randint(0,4)

itemPlayer4 = [0,0,0,0,0]
itemPlayer4[0] = random.randint(0,3)
itemPlayer4[1] = random.randint(0,3)
itemPlayer4[2] = random.randint(0,3)
itemPlayer4[3] = random.randint(0,4)
itemPlayer4[4] = random.randint(0,4)

posItemPlayer1 = [0,0,0,0]
posItemPlayer1[0]=random.randint(200,350)
posItemPlayer1[1]=random.randint(400,550)
posItemPlayer1[2]=random.randint(600,750)
posItemPlayer1[3]=random.randint(800,950)

posItemPlayer2 = [0,0,0,0]
posItemPlayer2[0]=random.randint(200,350)
posItemPlayer2[1]=random.randint(400,550)
posItemPlayer2[2]=random.randint(600,750)
posItemPlayer2[3]=random.randint(800,950)

posItemPlayer3 = [0,0,0,0]
posItemPlayer3[0]=random.randint(200,350)
posItemPlayer3[1]=random.randint(400,550)
posItemPlayer3[2]=random.randint(600,750)
posItemPlayer3[3]=random.randint(800,950)

posItemPlayer4 = [0,0,0,0]
posItemPlayer4[0]=random.randint(200,350)
posItemPlayer4[1]=random.randint(400,550)
posItemPlayer4[2]=random.randint(600,750)
posItemPlayer4[3]=random.randint(800,950)



bool1 = True
bool2 = True
bool3 = True
bool4 = True
bool5 = True
mapNumber = 0
setNumber = -1
playerNumber = -1
bet_amount = 0
sound = 1
ten=['','','','']
block_size = 20

#playerX = 100 
speedPlayer=[0,0,0,0]
posPlayer=[75,75,75,75]

start = False

listrank = [0,0,0,0]
rank = 0
cellMoney =''
cellHistoryMoney=''
cellHistoryRank=''
historyRank=''
historyMoney=''
historyRankList=[]
historyMoneyList=[]

cellHistoryUpDownMoney=''
historyUpDownMoneyList=[]

replay1 = -1


for i in range(0,4):
    speedPlayer[i]= random.randint(25,30)/10
tangToc=[False,False,False,False]
lamCham=[False,False,False,False]
choang=[False, False, False, False]
quayLai=[False, False, False, False]
demTangToc=[0,0,0,0]
demLamCham=[0,0,0,0]
demChoang=[0,0,0,0]
demQuayLai=[0,0,0,0]
temp=[0,0,0,0]
k=[0,0,0,0]

def text_to_button(msg, color, buttonx, buttony, buttonwidth, buttonheight, size = "small"):
	textSurf, textRect = text_objects(msg, color, size) 
	textRect.center = (buttonx + buttonwidth/2, buttony + buttonheight/2)
	screen.blit(textSurf, textRect)

def update_money():
	wb = openpyxl.load_workbook('Login.xlsx')
	sheet_login = wb['Sheet1']
	sheet_login[cellMoney].value = snakeMoney
	wb.close()
	wb.save('Login.xlsx')

def update_historyRankList():
    global historyRank, historyRankList
    wb = openpyxl.load_workbook('Login.xlsx')
    sheet_login = wb['Sheet1']
    historyRank = ".".join(historyRankList)
    sheet_login[cellHistoryRank].value = historyRank
    wb.close()
    wb.save('Login.xlsx')

def update_historyMoneyList():
    global historyMoney, historyMoneyList
    wb = openpyxl.load_workbook('Login.xlsx')
    sheet_login = wb['Sheet1']
    historyMoney = ".".join(historyMoneyList)
    sheet_login[cellHistoryMoney].value = historyMoney
    wb.close()
    wb.save('Login.xlsx')

def update_historyUpDownMoney():
    global historyUpDownMoney, historyUpDownMoneyList
    wb = openpyxl.load_workbook('Login.xlsx')
    sheet_login = wb['Sheet1']
    historyUpDownMoney = ".".join(historyUpDownMoneyList)
    sheet_login[cellHistoryUpDownMoney].value = historyUpDownMoney
    wb.close()
    wb.save('Login.xlsx')

def score(score):
	text = smallfont.render('Score: ' + str(score), True , RED)
	screen.blit(text, (30,30))

def snake(block_size, snakeList):
	for XnY in snakeList:
		pygame.draw.rect(screen, GREEN, (XnY[0],XnY[1],block_size,block_size))

def text_objects(text,color,size):
	if size == "small":
		textSurface = smallfont.render(text, True, color)
	elif size == "medium":
		textSurface = mediumfont.render(text, True, color)
	elif size == "large1":
		textSurface = largefont1.render(text, True, color)
	elif size == "large2":
		textSurface = largefont2.render(text, True, color)
	return textSurface, textSurface.get_rect()

def message_to_screen(msg, color,y=0, size = "small"):
	textSurf, textRect = text_objects(msg,color,size)
	textRect.center = (WINDOWWIDTH/2,(WINDOWHEIGHT/2)+y)
	screen.blit(textSurf, textRect)
def pause():
	paused = True
	while paused:
		for event in pygame.event.get():
			if event.type == pygame.QUIT:
				pygame.quit()
				quit()
			if event.type == pygame.KEYDOWN:
				if event.key == pygame.K_p:
					paused = False
		#screen.fill(GREY)
		message_to_screen('PAUSED',
							RED,
							y = -50,
							size = "large1")
		
		pygame.display.update()

		fpsClock.tick(FPS)
def get_value(cellname):
    filename='Login.xlsx'
    wb = openpyxl.load_workbook(filename)
    sheet_login = wb['Sheet1']
    wb.close()
    return str(sheet_login[cellname].value)
dem_so_tk= int(get_value('AA1')) 

def check_login(user,password):
    kt = False
    global snakeScore, username, snakeMoney, cellMoney
    global cellHistoryRank, historyRank, cellHistoryMoney, historyMoney, historyRankList, historyMoneyList, cellHistoryUpDownMoney, historyUpDownMoneyList,historyUpDownMoney
    for i in range (2,dem_so_tk+1):
        check_user=get_value('A'+str(i))
        check_password=get_value('B'+str(i))
        if user== check_user and password== check_password:
            kt = True
            cellMoney = 'C'+str(i)
            snakeMoney = int(get_value('C'+str(i)))
            cellHistoryRank='D'+str(i)
            historyRank = get_value(cellHistoryRank)
            historyRankList = historyRank.split(".")
            cellHistoryMoney='E'+str(i)
            historyMoney = get_value(cellHistoryMoney)
            historyMoneyList = historyMoney.split(".")
            cellHistoryUpDownMoney='F'+str(i)
            historyUpDownMoney = get_value(cellHistoryUpDownMoney)
            historyUpDownMoneyList = historyUpDownMoney.split(".")

            print(snakeScore)
            print(snakeMoney)
            username = get_value('A'+str(i))
            print(username)
            break
    if kt== True:
        return True
    else:
        return False

def up_down(x):
	running = True 
	while running:
		screen.fill((255,255,255))
		message_to_screen(str(x), RED, y = -200,size = "large1")
		message_to_screen("Your money: " + str(snakeMoney) + "$", RED, y = 100,size = "large1")
		for event in pygame.event.get():
			if event.type == pygame.MOUSEBUTTONDOWN:
				if event.button == 1:
					running = False



def login():
    font = pygame.font.Font(None, 32)
    clock = pygame.time.Clock()
    input_box1 = pygame.Rect(500, 300, 140, 32)
    input_box2 = pygame.Rect(500, 350, 140, 32)
    color_inactive = pygame.Color(BLUE)
    color_active = pygame.Color(YELLOW)
    color1 = color_inactive
    active1 = False
    color2 = color_inactive
    active2 = False
    text1 = ''
    text2 = ''
    text2_sao=''
    login_user=''
    login_password=''
    done = False

    while not done:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit() 
                quit()
                done = True
            if event.type == pygame.MOUSEBUTTONDOWN:
                # If the user clicked on the input_box rect.
                if input_box1.collidepoint(event.pos):
                    # Toggle the active variable.
                    active1 = not active1
                else:
                    active1 = False
                # Change the current color of the input box.
                color1 = color_active if active1 else color_inactive
            if event.type == pygame.KEYDOWN:
                if active1:
                    if event.key == pygame.K_BACKSPACE:
                        text1 = text1[:-1]
                    else:
                        text1 += event.unicode
            if event.type == pygame.MOUSEBUTTONDOWN:
                # If the user clicked on the input_box rect.
                if input_box2.collidepoint(event.pos):
                    # Toggle the active variable.
                    active2 = not active2
                else:
                    active2 = False
                # Change the current color of the input box.
                color2 = color_active if active2 else color_inactive
            if event.type == pygame.KEYDOWN:
                if active2:
                    if event.key == pygame.K_BACKSPACE:
                        text2 = text2[:-1]
                        text2_sao = text2_sao[:-1]
                    else:
                        text2 += event.unicode
                        text2_sao +='*'

        screen.fill((30, 30, 30))

        txt_surface1 = font.render(text1, True, color1)
        txt_surface2 = font.render(text2_sao, True, color2)
    
        width1 = max(200, txt_surface1.get_width()+10)
        width2 = max(200, txt_surface2.get_width()+10)
        input_box1.w = width1
        input_box2.w = width2

        screen.blit(txt_surface1, (input_box1.x+5, input_box1.y+5))
        screen.blit(txt_surface2, (input_box2.x+5, input_box2.y+5))

        pygame.draw.rect(screen, color1, input_box1, 2)
        pygame.draw.rect(screen, color2, input_box2, 2)

        login_Text=largefont3.render("LOGIN", True, RED)
        user=smallfont.render("Username ", True, BLUE)
        password=smallfont.render("Password ",True, BLUE)
        login_Boxtext=smallfont.render("Login", True, BLACK)
        register_Boxtext=smallfont.render("Create an account?", True, WHITE)
        screen.blit(login_Text,(450,100))
        screen.blit(user,(380,300))
        screen.blit(password,(380,350))
        login_Box = pygame.Surface((400,40))
        login_Box.fill(WHITE)
        login_Box.blit(login_Boxtext,(180,3))
        screen.blit(login_Box,(380,450))
        mouse_x, mouse_y = pygame.mouse.get_pos()
        register_Box = pygame.Surface((230,40))
        register_Box.fill((30,30,30))
        register_Box.blit(register_Boxtext,(0,4))
        screen.blit(register_Box,(500,500))
        mouse_x, mouse_y = pygame.mouse.get_pos()
        if event.type == pygame.MOUSEBUTTONDOWN:
            if event.button == 1 and (380 < mouse_x < 780) and (450 < mouse_y < 490):
                if check_login(text1,text2) == True:
                    done = True
                else:
                    wrong=smallfont.render("The username or password is incorrect!!!", True, RED)
                    screen.blit(wrong,(400,400))
            if event.button == 1 and (500 < mouse_x < 730) and (500 < mouse_y < 540):
                register()
        pygame.display.flip()
        clock.tick(30)
def history():
    running = True
    listRank_his=[0,0,0,0,0,0,0,0,0,0]
    listMoney_his=[0,0,0,0,0,0,0,0,0,0]
    listChange_his=[0,0,0,0,0,0,0,0,0,0]
    h=0

    for i in range(0,8):
        if int(historyRankList[i])>0:
        	h=h+1
    while running:    
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit() 
                quit()
            history_tab = pygame.Surface((866,500))
            history_tab.fill(WHITE)
            history_Text=largefont4.render("HISTORY", True, BLACK)
            history_tab.blit(history_Text,(75,75))
            user=smallfont.render(username, True, BLACK)
            history_tab.blit(user,(10,10))
            money1=smallfont.render("Your money: " + str(snakeMoney) + "$" , True, RED)
            history_tab.blit(money1,(500,10))
            rank_Text=smallfont.render("Rank", True, BLACK)
            history_tab.blit(rank_Text,(150,150))
            money_Text=smallfont.render("Money", True, BLACK)
            history_tab.blit(money_Text,(450,150))
            change_Text=smallfont.render("Change", True, BLACK)
            history_tab.blit(change_Text,(300,150))

            for i in range(0,h):
                listRank_his[i]=smallfont.render(str(historyRankList[i]),True,BLACK)
                history_tab.blit(listRank_his[i],(150,200+i*50))
                listMoney_his[i]=smallfont.render(str(historyMoneyList[i]),True,BLACK)
                history_tab.blit(listMoney_his[i],(450,200+i*50))
                listChange_his[i]=smallfont.render(str(historyUpDownMoneyList[i]),True,BLACK)
                history_tab.blit(listChange_his[i],(300,200+i*50))
            mouse_x, mouse_y = pygame.mouse.get_pos()
 
            if event.type == pygame.MOUSEBUTTONDOWN:
                if 866 > mouse_x-217 > 826 and 433 > mouse_y-108 > 0 and event.button == 1:
                    running = False
            quit_history = smallfont.render("X", True, BLACK)
            pygame.draw.line(history_tab, BLACK, (0,40), (866,40), 3)
            history_tab.blit(quit_history,(836,5))  
            screen.blit(history_tab,(217,108))
            pygame.display.flip() 
def register():

    font = pygame.font.Font(None, 32)
    register = True
    register_user = ''
    register_password = ''
    register_password_sao = ''
    register_confirm = ''
    register_confirm_sao = ''
    color_inactive = pygame.Color(BLACK)
    color_active = pygame.Color(RED)
    color3 = color_inactive
    active3 = False
    color4 = color_inactive
    active4 = False
    color5 = color_inactive
    active5 = False
    done = False
    filename='Login.xlsx'
    wb = openpyxl.load_workbook(filename)
    sheet_login = wb['Sheet1']
    global dem_so_tk

    while register:    
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit() 
                quit()
            register_tab = pygame.Surface((866,500))
            register_tab.fill(WHITE)
            register_Text=largefont4.render("CREATE YOUR ACCOUNT", True, BLACK)
            register_tab.blit(register_Text,(75,75))
            register_input_user = pygame.Rect(400, 200, 140, 32)
            register_input_password = pygame.Rect(400, 250, 140, 32)
            register_input_confirm = pygame.Rect(400, 300, 140, 32)
            user=smallfont.render("Username", True, BLACK)
            password= smallfont.render("Password",True, BLACK)
            confirm= smallfont.render("Repeat password",True, BLACK)
            register_tab.blit(user,(190,200))
            register_tab.blit(password,(190,250))
            register_tab.blit(confirm,(190,300))
            mouse_x, mouse_y = pygame.mouse.get_pos()
            if event.type == pygame.MOUSEBUTTONDOWN:
                if 866 > mouse_x-217 > 826 and 433 > mouse_y-108 > 0 and event.button == 1:
                    register = False
            quit_register = smallfont.render("X", True, BLACK)
            pygame.draw.line(register_tab, BLACK, (0,40), (866,40), 3)
            register_tab.blit(quit_register,(836,5))
            if event.type == pygame.MOUSEBUTTONDOWN:
                # If the user clicked on the input_box rect.
                if event.button == 1 and 350<mouse_x-217<350+200 and 200<mouse_y-108<200+32:
                    # Toggle the active variable.
                    active3 = True
                else:
                    active3 = False
                # Change the current color of the input box.
            color3 = color_active if active3 else color_inactive
            if active3:
                if event.type == pygame.KEYDOWN:
                    if event.key == pygame.K_BACKSPACE:
                        register_user = register_user[:-1]
                    else:
                        register_user += event.unicode
                    
                    
            if event.type == pygame.MOUSEBUTTONDOWN:
                # If the user clicked on the input_box rect.
                if event.button == 1 and 350<mouse_x-217<350+200 and 250<mouse_y-108 <250+32:
                    # Toggle the active variable.
                    active4 = True
                else:
                    active4 = False
                # Change the current color of the input box.
                color4 = color_active if active4 else color_inactive
            if active4:
                if event.type == pygame.KEYDOWN:
                        if event.key == pygame.K_BACKSPACE:
                            register_password = register_password[:-1]
                            register_password_sao = register_password_sao[:-1]
                        else:
                            register_password += event.unicode
                            register_password_sao +='*'

            if event.type == pygame.MOUSEBUTTONDOWN:
                # If the user clicked on the input_box rect.
                if event.button == 1 and 350<mouse_x-217<350+200 and 300<mouse_y-108 <300+32:
                    # Toggle the active variable.
                    active5 = True
                else:
                    active5 = False
                # Change the current color of the input box.
                color5 = color_active if active5 else color_inactive
            if active5:
                if event.type == pygame.KEYDOWN:
                        if event.key == pygame.K_BACKSPACE:
                            register_confirm = register_confirm[:-1]
                            register_confirm_sao = register_confirm_sao[:-1]
                        else:
                            register_confirm += event.unicode
                            register_confirm_sao +='*'

            txt_register_1 = font.render(register_user, True, color3)
            txt_register_2 = font.render(register_password_sao, True, color4)
            txt_register_3 = font.render(register_confirm_sao, True, color5)
    
            width1 = max(200, txt_register_1.get_width()+10)
            width2 = max(200, txt_register_2.get_width()+10)
            width3 = max(200, txt_register_3.get_width()+10)
            register_input_user.w = width1
            register_input_password.w = width2
            register_input_confirm.w = width3
        
            pygame.draw.rect(register_tab, color3, register_input_user, 2)
            pygame.draw.rect(register_tab, color4, register_input_password, 2)
            pygame.draw.rect(register_tab, color5, register_input_confirm, 2)

            register_tab.blit(txt_register_1, (register_input_user.x+5, register_input_user.y+5))
            register_tab.blit(txt_register_2, (register_input_password.x+5, register_input_password.y+5))
            register_tab.blit(txt_register_3, (register_input_confirm.x+5, register_input_confirm.y+5))

            register_Boxtext=smallfont.render("REGISTER", True, WHITE)
            register_Box = pygame.Surface((410,40))
            register_Box.fill(BLACK)
            register_Box.blit(register_Boxtext,(155,3))
            register_tab.blit(register_Box, (190,350))
            if event.type == pygame.MOUSEBUTTONDOWN:
                if event.button == 1 and 190<mouse_x-217<190+410 and 350<mouse_y-108 <350+40 and done == False:
                    if register_confirm == register_password:
                        dem_so_tk +=1
                        conclusion= smallfont1.render("Registered successfully!!!",True, RED)
                        register_tab.blit(conclusion,(190,400))
                        print("Registered successfully!!!")
                        done = True
                        sheet_login['AA1'].value = str(dem_so_tk)
                        sheet_login['A'+str(dem_so_tk)].value = register_user
                        sheet_login['B'+str(dem_so_tk)].value = register_password
                        sheet_login['C'+str(dem_so_tk)].value = '0'
                        sheet_login['D'+str(dem_so_tk)].value = '0.0.0.0.0.0.0.0'
                        sheet_login['E'+str(dem_so_tk)].value = '0.0.0.0.0.0.0.0'
                        sheet_login['F'+str(dem_so_tk)].value = '0.0.0.0.0.0.0.0'

                        print(dem_so_tk-1)
                        wb.close()
                        wb.save(filename)
                    else:
                        conclusion= smallfont1.render("Passwords do not match!!!",True, RED)
                        register_tab.blit(conclusion,(300,400))  

            screen.blit(register_tab,(217,108))
            pygame.display.flip() 

def replay():
    global replay1
    running = True
    while running:    
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
                pygame.quit() 
                quit()
            message_to_screen('Play again?', RED, y = -200,size = "large1")
            replay_tab = pygame.Surface((700,200),SRCALPHA)
            #replay_tab.fill(YELLOW)
            quit = pygame.image.load('quit.png')
            replay = pygame.image.load('replay.png')
            replay_tab.blit(replay,(33,25))
            replay_tab.blit(quit,(66+150+100,25))
            money2=smallfont.render("Your money: " + str(snakeMoney) + "$" , True, WHITE)
            screen.blit(money2,(600,20))
            screen.blit(replay_tab,(450,300))
            mouse_x, mouse_y = pygame.mouse.get_pos()
            if event.type == pygame.MOUSEBUTTONDOWN:
                if 483 < mouse_x < 483+650 and 250+75 < mouse_y < 400+75 and event.button == 1:
                    replay1 = 1
                    running = False
                if 483+183+100< mouse_x < 100+483+183+650 and 250+75 < mouse_y <400+75 and event.button == 1:
                    replay1 = 2
                    running = False
            pygame.display.flip() 
def startButton():
    global bool1, start
    while start==False:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit() 
                quit()
                done = True
            mouse_x, mouse_y = pygame.mouse.get_pos()
            if event.type == pygame.MOUSEBUTTONDOWN:
                if event.button == 1 and (575 < mouse_x < 575+150) and (400 < mouse_y < 475):
                	bool1 = False
                	start = True
                if event.button == 1 and (575 < mouse_x < 575+150) and (500 < mouse_y < 575):
                	history()
            
              
        screen.blit(backgroundStart, (0,0))
        message_to_screen('MORTAL RACE', RED, y = -150,size = "large1")

        login_Boxtext=smallfont.render("START", True, BLACK)
        login_Box = pygame.Surface((150,75))
        mouse_x, mouse_y = pygame.mouse.get_pos()

        if (575 < mouse_x < 575+150) and (400 < mouse_y < 475):
            pygame.draw.rect(login_Box, WHITE, (0 , 0, 150, 75))
        else:
            pygame.draw.rect(login_Box, YELLOW, (0 , 0, 150, 75))
        #login_Box.fill(WHITE)
        login_Box.blit(login_Boxtext,(30,15))
        screen.blit(login_Box,(575,400))

        history_Boxtext=smallfont.render("HISTORY", True, BLACK)
        history_Box = pygame.Surface((150,75))
        pygame.draw.rect(history_Box, BLUE, (0 , 0, 150, 150))
        #login_Box.fill(WHITE)
        history_Box.blit(history_Boxtext,(15,15))
        screen.blit(history_Box,(575,500))
        	
        pygame.display.flip()
def chooseMapButton(inactive_color, active_color):
	global bool2
	global mapNumber, chooseMap
	if bool2 == True:
		message_to_screen(str(snakeMoney)+'$',
							RED,
							y = 150,
							size = "small")
		screen.blit(mapstart,(0,0))
		mouse_x, mouse_y = pygame.mouse.get_pos()
		if mouse_x < 950 and mouse_x > 350 and mouse_y < 300 and mouse_y > 200:
			pygame.draw.rect(screen, active_color, (350-2, 200-2, 603, 103),2)
		elif mouse_x < 950 and mouse_x > 350 and mouse_y < 450 and mouse_y > 350:
			pygame.draw.rect(screen, active_color, (350-2, 350-2, 603, 103),2)
		elif mouse_x < 950 and mouse_x > 350 and mouse_y < 600 and mouse_y > 500:
			pygame.draw.rect(screen, active_color, (350-2, 500-2, 603, 103),2)

		message_to_screen('CHOOSE MAP', 
							BLACK, 
							y=-200, 
							size = "large1")
		for event in pygame.event.get():
			if event.type == pygame.QUIT:
				pygame.quit()
				quit()
			if event.type == pygame.MOUSEBUTTONDOWN:
				if event.button == 1 and mouse_x < 950 and mouse_x > 350 and mouse_y < 300 and mouse_y > 200:
					bool2 = False
					mapNumber = 0
				elif event.button == 1 and mouse_x < 950 and mouse_x > 350 and mouse_y < 450 and mouse_y > 350:
					bool2= False
					mapNumber = 1
				elif event.button == 1 and mouse_x < 950 and mouse_x > 350 and mouse_y < 600 and mouse_y > 500:
					bool2= False
					mapNumber = 2
		screen.blit(theme1,(350,200))
		screen.blit(theme2,(350,350))
		screen.blit(theme3,(350,500))
	#text_to_button(text,BLACK,x,y,width,height)
def chooseSet(inactive_color, active_color):
	global bool3
	global setNumber
	
	if bool3 == True:
		screen.blit(backgroundStart,(0,0))
		mouse_x, mouse_y = pygame.mouse.get_pos()
		if 100 < mouse_x < 600 and 250< mouse_y < 400:
			pygame.draw.rect(screen, YELLOW, (100-2, 250-2, 508, 158),3)
		elif 100 < mouse_x < 600 and 450< mouse_y < 600:
			pygame.draw.rect(screen, YELLOW, (100-2, 450-2, 508, 158),3)
		elif 700 < mouse_x < 1200 and 250< mouse_y < 400:
			pygame.draw.rect(screen, YELLOW, (700-2, 250-2, 508, 158),3)
		elif 700 < mouse_x < 1200 and 450< mouse_y < 600:
			pygame.draw.rect(screen, YELLOW, (700-2, 450-2, 506, 158),3)

		message_to_screen('CHOOSE SET', 
						RED, 
						y=-200, 
						size = "large1")
		for event in pygame.event.get():
			if event.type == pygame.QUIT:
					pygame.quit()
					quit()
			if event.type == pygame.MOUSEBUTTONDOWN:
				if event.button == 1 and 100 < mouse_x < 600 and 250 < mouse_y < 400:
					bool3 = False
					setNumber = 0
				elif event.button == 1 and 100 < mouse_x < 600 and 450< mouse_y < 600:
					bool3= False
					setNumber = 1
				elif event.button == 1 and 700 < mouse_x < 1200 and 250< mouse_y < 400:
					bool3= False
					setNumber = 2
				elif event.button == 1 and 700 < mouse_x < 1200 and 450< mouse_y < 600:
					bool3= False
					setNumber = 3

		screen.blit(setList[0],(100,250))
		screen.blit(setList[1],(100,450))
		screen.blit(setList[2],(700,250))
		screen.blit(setList[3],(700,450))

def choosePlayer():
	global bool4, playerNumber,ten
	if bool4 == True:
		screen.blit(playerChoose[setNumber],(0,0))
		mouse_x, mouse_y = pygame.mouse.get_pos()
		if 0 < mouse_x < 325 and 0< mouse_y < 650:
			pygame.draw.rect(screen, RED, (0, 0, 325, 650),4)
		elif 325 < mouse_x < 650 and 0< mouse_y < 650:
			pygame.draw.rect(screen, RED, (325, 0, 325, 650),4)
		elif 650 < mouse_x < 975 and 0< mouse_y < 650:
			pygame.draw.rect(screen, RED, (650, 0, 325, 650),4)
		elif 975 < mouse_x < 1300 and 0< mouse_y < 650:
			pygame.draw.rect(screen, RED, (975, 0, 325, 650),4)

		message_to_screen('CHOOSE PLAYER', 
						RED, 
						y=200, 
						size = "large1")
		for event in pygame.event.get():
			if event.type == pygame.QUIT:
				pygame.quit()
				quit()
			if event.type == pygame.MOUSEBUTTONDOWN:
				if event.button == 1 and 0 < mouse_x < 325 and 0< mouse_y < 650:
					bool4 = False
					playerNumber = 0
				elif event.button == 1 and 325 < mouse_x < 650 and 0< mouse_y < 650:
					bool4= False
					playerNumber = 1
				elif event.button == 1 and 650 < mouse_x < 975 and 0< mouse_y < 650:
					bool4= False
					playerNumber = 2
				elif event.button == 1 and 975 < mouse_x < 1300 and 0< mouse_y < 650:
					bool4= False
					playerNumber = 3
		bot = 1 
		for i in range (0,4):
			if i == playerNumber:
				ten[i] = username
			else:
				ten[i]= "BOT"+str(bot)
				bot+=1
def chonMucCuoc(active_color, inactive_color):
	global bool5, bet_amount
	if bool5 == True:
		screen.blit(choose_bet,(0,0))
		mouse_x, mouse_y = pygame.mouse.get_pos()
		message_to_screen('CHOOSE BET AMOUNT',
							BLACK, 
							y=-200, 
							size = "large1")
		for event in pygame.event.get():
			if event.type == pygame.QUIT:
				pygame.quit()
				quit()
			if event.type == pygame.MOUSEBUTTONDOWN:
				if event.button == 1 and 463< mouse_x <837 and 257<mouse_y<257+85:
					bool5 = False
					bet_amount = 1000
				elif event.button == 1 and 463< mouse_x <837 and 381<mouse_y<466:
					bool5= False
					bet_amount = 3000
				elif event.button == 1 and 463< mouse_x <837 and 505<mouse_y<590:
					bool5= False
					bet_amount = 5000
				print(bet_amount)

	#text_to_button(text,BLACK,x,y,width,height)

def game_intro1():
	intro = True
	while intro:
		for event in pygame.event.get():
			if event.type == pygame.QUIT:
				pygame.quit()
				quit()
			if event.type == pygame.KEYDOWN:
				if event.key == pygame.K_s:
					intro = False
				if event.key == pygame.K_q:
					pygame.quit()
					quit()
		screen.blit(background[3],(0,0))
		text = smallfont.render(username, True, RED)
		screen.blit(text,(600,30))
		text1 = smallfont.render('Money: ' +str(snakeMoney) + '$', True, RED)
		screen.blit(text1,(1100,30))
		message_to_screen('Welcome to Slither',
							GREEN,
							y = -100,
							size = "large2")
		message_to_screen('*This is a mini game to earn the minimum amount to bet*',
							RED,
							y = 300,
							size = "small")
		message_to_screen('The objective of the game is to eat red apples',
							BLACK,
							y = -30,
							size = "small")
		message_to_screen('The more apples you eat, the longer you get',
							BLACK,
							y = 20,
							size = "small")
		message_to_screen('If you run into yourself, or the edges, you die!!!',
							BLACK,
							y = 70,
							size = "small")
		message_to_screen('Press S to start or Q to quit',
							BLACK,
							y = 150,
							size = "small")
		pygame.display.update()		

def game_intro2():
	global bool1, bool2, bool3
	intro = True
	pygame.mixer.music.load("y2mate.com - Old School Cypher Beat Eternal _ Cypher Rap Beat Instrumental (Prod. Chuki Beats).mp3")
	pygame.mixer.music.play(-1)
	while intro:
		for event in pygame.event.get():
			if event.type == pygame.QUIT:
				pygame.quit()
				quit()
		startButton()
		if bool1 == False:
			chooseMapButton(WHITE,BLACK)
			if bool2 == False:
				chooseSet(WHITE,BLACK)
				if bool3 == False:
					for i in range(0,4):
						player[i]= character_move[setNumber*4+i]
					choosePlayer()
					if bool4 == False:
						chonMucCuoc(BLACK, WHITE)
						if bool5 == False:
							pygame.mixer.music.stop()
							pygame.mixer.music.load("nhac nen.mp3")
							pygame.mixer.music.play(-1)
							intro = False

		pygame.display.update()


def MainGame():
	global rank,k
	global listrank
	global sound, tangtoc, lamcham	, demTangtoc, demLamcham, temp, money_change
	for i in range(0,4):	
		if posPlayer[i]<1220:
			posPlayer[i] += speedPlayer[i]
		if posPlayer[i]>=1220 and posPlayer[i]!=1230 :
			posPlayer[i]=1230
			rank+=1
			listrank[i] = rank
			player[i]=character[setNumber*4+i]
		screen.blit(player[i],(posPlayer[i],30+162*i))

		#ăn bùa
		if posPlayer[i]<1220:
			if k[i]<=3 and i==0:
				if posPlayer[i]+25< posItemPlayer1[k[i]] :
					screen.blit(item[itemPlayer1[k[i]]], (posItemPlayer1[k[i]],30))
				elif posPlayer[i]+25>= posItemPlayer1[k[i]]:
					if itemPlayer1[k[i]] == 2:
						quayLai[i]=True
						speedPlayer[i]*=-1
						demQuayLai[i]=0
					if itemPlayer1[k[i]] == 4:
						posPlayer[i] =1221
					if itemPlayer1[k[i]] == 0:
						tangToc[i]=True
						demTangToc[i]=0
						speedPlayer[i]*=1.5
					if itemPlayer1[k[i]] == 1:
						lamCham[i]=True
						demLamCham[i]=0
						speedPlayer[i]*=0.5
					if itemPlayer1[k[i]] == 3:
						choang[i]=True
						demChoang[i]=0
						temp[i] = speedPlayer[i] 
						speedPlayer[i] = 0
					k[i]+=1
			if k[i]<=3 and i==1:
				if posPlayer[i]+25< posItemPlayer2[k[i]] :
					screen.blit(item[itemPlayer2[k[i]]], (posItemPlayer2[k[i]],192))
				elif posPlayer[i]+25>= posItemPlayer2[k[i]]:
					if itemPlayer2[k[i]] == 2:
						quayLai[i]=True
						speedPlayer[i]*=-1
						demQuayLai[i]=0
					if itemPlayer2[k[i]] == 4:
						posPlayer[i] =1221
					if itemPlayer2[k[i]] == 0:
						tangToc[i]=True
						demTangToc[i]=0
						speedPlayer[i]*=1.5
					if itemPlayer2[k[i]] == 1:
						lamCham[i]=True
						demLamCham[i]=0
						speedPlayer[i]*=0.5
					if itemPlayer2[k[i]] == 3:
						choang[i]=True
						demChoang[i]=0
						temp[i] = speedPlayer[i] 
						speedPlayer[i] = 0
					k[i]+=1
			if k[i]<=3 and i==2:
				if posPlayer[i]+25< posItemPlayer3[k[i]] :
					screen.blit(item[itemPlayer3[k[i]]], (posItemPlayer3[k[i]],354))
				elif posPlayer[i]+25>= posItemPlayer3[k[i]]:
					if itemPlayer3[k[i]] == 2:
						quayLai[i]=True
						speedPlayer[i]*=-1
						demQuayLai[i]=0
					if itemPlayer3[k[i]] == 4:
						posPlayer[i] =1221
					if itemPlayer3[k[i]] == 0:
						tangToc[i]=True
						demTangToc[i]=0
						speedPlayer[i]*=1.5
					if itemPlayer3[k[i]] == 1:
						lamCham[i]=True
						demLamCham[i]=0
						speedPlayer[i]*=0.5
					if itemPlayer3[k[i]] == 3:
						choang[i]=True
						demChoang[i]=0
						temp[i] = speedPlayer[i] 
						speedPlayer[i] = 0
					k[i]+=1
			if k[i]<=3 and i==3:
				if posPlayer[i]+25< posItemPlayer4[k[i]] :
					screen.blit(item[itemPlayer4[k[i]]], (posItemPlayer4[k[i]],516))
				elif posPlayer[i]+25>= posItemPlayer4[k[i]]:
					if itemPlayer4[k[i]] == 2:
						quayLai[i]=True
						speedPlayer[i]*=-1
						demQuayLai[i]=0
					if itemPlayer4[k[i]] == 4:
						posPlayer[i] =1221
					if itemPlayer4[k[i]] == 0:
						tangToc[i]=True
						demTangToc[i]=0
						speedPlayer[i]*=1.5
					if itemPlayer4[k[i]] == 1:
						lamCham[i]=True
						demLamCham[i]=0
						speedPlayer[i]*=0.5
					if itemPlayer4[k[i]] == 3:
						choang[i]=True
						demChoang[i]=0
						temp[i] = speedPlayer[i] 
						speedPlayer[i] = 0
					k[i]+=1
		if quayLai[i]==True:
			demQuayLai[i]+=1
			player[i]= character_move_1[setNumber*4+i]
			screen.blit(quaylai,(posPlayer[i]+50,162*i))
			if demQuayLai[i]>35:
				player[i]= character_move[setNumber*4+i] 
				speedPlayer[i]*=-1
				quayLai[i]=False
		if tangToc[i]==True:
			demTangToc[i]+=1
			screen.blit(nhanh,(posPlayer[i]+60,162*i))
			if demTangToc[i]>25:
				speedPlayer[i]/=1.5
				tangToc[i]=False
		if lamCham[i]==True:
			demLamCham[i]+=1
			screen.blit(cham,(posPlayer[i]+50,162*i))
			if demLamCham[i]>25:
				speedPlayer[i]/=0.75
				lamCham[i]=False
		if choang[i]==True:
			demChoang[i]+=1
			screen.blit(Choang,(posPlayer[i],162*i))
			player[i]= character[setNumber*4+i]
			if demChoang[i]>35:
				speedPlayer[i] = temp[i]
				player[i]= character_move[setNumber*4+i]
				choang[i]=False

	for i in range(0,4):
		if i == playerNumber:
			text = font.render(ten[i], True, RED)
			screen.blit(text,(3,60+160*i))
		else: 
			text = font.render(ten[i], True, WHITE)
			screen.blit(text,(3,60+160*i))
	if rank==4:
		if listrank[playerNumber] == 1:
			message_to_screen('WIN', RED, y = -200,size = "large1")
			money_change = bet_amount
		else:
			message_to_screen('LOSE', RED, y = -200,size = "large1")
			money_change = bet_amount * -1
		print(money_change)
		for i in range(1,5):
			for j in range(0,4):
				if listrank[j] == i:
					rankSurface=mediumfont.render('Rank #' + str(i) + ' is ' + ten[j] ,True,WHITE)
					screen.blit(rankSurface,(500,i*100+100))
	#Nut bat/tat am thanh:
	if sound==1:
		screen.blit(bat,(1070,563))
	if sound==0:
		screen.blit(tat,(1070,563))
	mouse_x, mouse_y = pygame.mouse.get_pos()
	for event in pygame.event.get():
		if event.type == pygame.QUIT:
			pygame.quit()
			quit()
		if event.type == pygame.MOUSEBUTTONDOWN:
			if event.button == 1 and sound==1 and mouse_x>1070 and mouse_x<1070+70 and mouse_y>563 and mouse_y<563+70:
				sound=0
				pygame.mixer.music.set_volume(0.0)
			
			elif event.button == 1 and sound==0 and mouse_x>1070 and mouse_x<1070+70 and mouse_y>563 and mouse_y<563+70:
				sound=1
				pygame.mixer.music.set_volume(1.0)
		if event.type == pygame.KEYDOWN:
				if event.key == pygame.K_p:
					pause()
	pygame.display.update()

def gameLoop2():
	
	running = True
	gameOver = False
	lead_x = 650
	lead_y = 400
	lead_x_change = 0
	lead_y_change = 0
	snakeList = []
	snakeLength = 1
	global snakeMoney
	global snakeScore

	
	randAppleX = round(random.randrange(40, 1255)/10.0)*10.0
	randAppleY = round(random.randrange(40, 600)/10.0)*10.0
	pre=0 # biến lưu thao tác trước đó của rắn lần lượt tương ứng: trái, phải, lên, xuống: -1, 1, 2, -2
	while running:
		while gameOver == True:
			
			message_to_screen('GAME OVER',
								RED,
								y = -70,
								size = "large2")
			message_to_screen('Press A to play again or Q to quit',
								BLACK,
								y = 0,
								size = "small")
			pygame.display.update()
			if gameOver == True and snakeMoney >= 6000:
				#pause()
				game_intro2()

			for event in pygame.event.get():
				if event.type == pygame.QUIT:
					running = False
					pygame.quit()
					quit()
				if event.type == pygame.KEYDOWN:
					if event.key == pygame.K_a:
						gameLoop2()
						gameOver = False
					if event.key == pygame.K_q:
						pygame,quit()
						quit()

		for event in pygame.event.get():
			if event.type == pygame.QUIT:
				running = False
				pygame.quit() 
				quit()
			if event.type == pygame.KEYDOWN:
				if event.key == pygame.K_LEFT and pre!=1:
						lead_x_change = -10
						lead_y_change = 0
						pre=-1
				elif event.key == pygame.K_RIGHT and pre!= -1:
						lead_x_change = 10
						lead_y_change = 0
						pre=1
				elif event.key == pygame.K_UP and pre!= -2:
						lead_y_change = -10
						lead_x_change = 0
						pre=2
				elif event.key == pygame.K_DOWN and pre!= 2:
						lead_y_change = 10
						lead_x_change = 0
						pre=-2
				elif event.key == pygame.K_p:
						pause()
		lead_x += lead_x_change
		lead_y += lead_y_change
		if lead_x >=  1260 or lead_x <= 25 or lead_y >= 605 or lead_y <= 25:
			gameOver = True
			if lead_x >=  1255:
				lead_x =  1255
			if lead_x <= 25:
				lead_x = 25
			if lead_y >= 600:
				lead_y = 600
			if lead_y <= 25:
				lead_y = 23
		if gameOver == True and snakeMoney >= 6000:
			running =False
		
		
		screen.blit(background[4],(0,0))
		screen.blit(appleimg, (randAppleX,randAppleY))
		pygame.draw.rect(screen, BLACK, (20, 20, 1255, 605), 8)
		
		snakeHead = []
		snakeHead.append(lead_x)
		snakeHead.append(lead_y)
		snakeList.append(snakeHead)
		if len(snakeList) > snakeLength:
			del snakeList[0]
		for eachSegment in snakeList[:-1]:
			if eachSegment == snakeHead:
				gameOver = True
				running = False

		

		score(snakeScore)
		text = smallfont.render('Money: ' +str(snakeMoney) + '$', True, RED)
		screen.blit(text,(1100,30))
		snake(block_size, snakeList)
		pygame.display.update()	
		if (lead_x-15 <= randAppleX <= lead_x +15) and (lead_y-15  <= randAppleY <= lead_y+15):
			randAppleX = round(random.randrange(40, 1255)/10.0)*10.0
			randAppleY = round(random.randrange(40, 600)/10.0)*10.0
			snakeLength += 4
			snakeScore += 1
			snakeMoney += 1000
		fpsClock.tick(50)
	update_money()

	pygame.display.update()

def gameLoop1():
	global player, itemPlayer1, itemPlayer1, itemPlayer2, itemPlayer3, itemPlayer4, posItemPlayer1, posItemPlayer2, posItemPlayer3, posItemPlayer4, snakeMoney
	global bool1, bool2, bool3, bool4, start, setNumber,playerNumber, mapNumber, speedPlayer, posPlayer, listrank, rank, speedPlayer, k, replay1, money_change

	if snakeMoney<6000:
		game_intro1()
	while (snakeMoney<6000):
		gameLoop2()
	game_intro2()
	
	running=True
	kt =0
	while running:
		screen.fill((0,0,0))
		screen.blit(background[mapNumber],(0,0))
		if rank<4:
			MainGame()
		if rank==4:
			pygame.mixer.music.stop()
			if listrank[playerNumber] == 1 and kt==0:
				pygame.mixer.music.load("chuc mung.mp3")
				kt+=1
				pygame.mixer.music.play(-1)
				pygame.mixer.music.stop()

			else:
				pygame.mixer.music.load("that vong.mp3")
				kt+=1
				pygame.mixer.music.play(-1)
				pygame.mixer.music.stop()
			for event in pygame.event.get():
				if event.type == pygame.QUIT:
				    running = False
				if event.type == pygame.MOUSEBUTTONDOWN:
					if event.button == 1:
						snakeMoney += money_change
						for i in range(0,7):
							historyRankList[7-i]=historyRankList[7-1-i]
						historyRankList[0] = str(listrank[playerNumber])
						for i in range(0,7):
							historyMoneyList[7-i]=historyMoneyList[7-1-i]
						historyMoneyList[0] = str(snakeMoney)
						for i in range(0,7):
							historyUpDownMoneyList[7-i]=historyUpDownMoneyList[7-1-i]
						historyUpDownMoneyList[0] = str(money_change)
						update_money()
						update_historyMoneyList()
						update_historyRankList()
						update_historyUpDownMoney()

						running = False
						player=[0,0,0,0]

						itemPlayer1 = [0,0,0,0,0]
						itemPlayer1[0] = random.randint(0,3)
						itemPlayer1[1] = random.randint(0,3)
						itemPlayer1[2] = random.randint(0,3)
						itemPlayer1[3] = random.randint(0,4)
						itemPlayer1[4] = random.randint(0,4)

						itemPlayer2 = [0,0,0,0,0]
						itemPlayer2[0] = random.randint(0,3)
						itemPlayer2[1] = random.randint(0,3)
						itemPlayer2[2] = random.randint(0,3)
						itemPlayer2[3] = random.randint(0,4)
						itemPlayer2[4] = random.randint(0,4)

						itemPlayer3 = [0,0,0,0,0]
						itemPlayer3[0] = random.randint(0,3)
						itemPlayer3[1] = random.randint(0,3)
						itemPlayer3[2] = random.randint(0,3)
						itemPlayer3[3] = random.randint(0,4)
						itemPlayer3[4] = random.randint(0,4)
	
						itemPlayer4 = [0,0,0,0,0]
						itemPlayer4[0] = random.randint(0,3)
						itemPlayer4[1] = random.randint(0,3)
						itemPlayer4[2] = random.randint(0,3)
						itemPlayer4[3] = random.randint(0,4)
						itemPlayer4[4] = random.randint(0,4)
	
						posItemPlayer1 = [0,0,0,0]
						posItemPlayer1[0]=random.randint(200,350)
						posItemPlayer1[1]=random.randint(400,550)
						posItemPlayer1[2]=random.randint(600,750)
						posItemPlayer1[3]=random.randint(800,950)
	
						posItemPlayer2 = [0,0,0,0]
						posItemPlayer2[0]=random.randint(200,350)
						posItemPlayer2[1]=random.randint(400,550)
						posItemPlayer2[2]=random.randint(600,750)	
						posItemPlayer2[3]=random.randint(800,950)
	
						posItemPlayer3 = [0,0,0,0]
						posItemPlayer3[0]=random.randint(200,350)
						posItemPlayer3[1]=random.randint(400,550)
						posItemPlayer3[2]=random.randint(600,750)
						posItemPlayer3[3]=random.randint(800,950)
	
						posItemPlayer4 = [0,0,0,0]
						posItemPlayer4[0]=random.randint(200,350)
						posItemPlayer4[1]=random.randint(400,550)
						posItemPlayer4[2]=random.randint(600,750)
						posItemPlayer4[3]=random.randint(800,950)



						bool1 = True
						bool2 = True
						bool3 = True
						bool4 = True
						mapNumber = 0
						setNumber = -1
						playerNumber = -1

						speedPlayer=[0,0,0,0]
						posPlayer=[75,75,75,75]

						start = False
						listrank = [0,0,0,0]
						rank = 0
						for i in range(0,4):
							speedPlayer[i]= random.randint(25,30)/10
						tangToc=[False,False,False,False]
						lamCham=[False,False,False,False]
						choang=[False, False, False, False]
						quayLai=[False, False, False, False]
						demTangToc=[0,0,0,0]
						demLamCham=[0,0,0,0]
						demChoang=[0,0,0,0]
						demQuayLai=[0,0,0,0]
						temp=[0,0,0,0]
						k=[0,0,0,0]
						replay1 = -1
						print(money_change)
						replay()
						if replay1 == 1:
							gameLoop1()
						elif replay1 == 2:
							running = False
						
			if listrank[playerNumber] == 1:
				message_to_screen('WIN', RED, y = -200,size = "large1")
			else:
				message_to_screen('LOSE', RED, y = -200,size = "large1")
			for i in range(1,5):
				for j in range(0,4):
					if listrank[j] == i:
						rankSurface=mediumfont.render('Rank #' + str(i) + ' is ' + ten[j] ,True,WHITE)
						screen.blit(rankSurface,(500,i*100+100))
						screen.blit(player[j],(posPlayer[j],30+162*j))
		for event in pygame.event.get():
			if event.type == pygame.QUIT:
				    running = False

		pygame.display.update()
		fpsClock.tick(FPS)


	pygame.display.update()
	'''pygame.quit()
	quit()'''     
login()
gameLoop1()


pygame.quit()

